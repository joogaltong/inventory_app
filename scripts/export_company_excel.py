#!/usr/bin/env python3
import base64
import io
import json
import re
import sys
from difflib import SequenceMatcher
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

EXPORT_ZONES = ["피트니스", "게스트하우스", "계단실"]


def normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", "", text)
    text = text.replace("(", "").replace(")", "")
    return text


def to_number(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def cell_text(ws, row: int, col: int) -> str:
    value = ws.cell(row=row, column=col).value
    return str(value).strip() if value is not None else ""


def find_4danji_sheet_name(workbook) -> str:
    if "4단지" in workbook.sheetnames:
        return "4단지"
    for name in workbook.sheetnames:
        if str(name).strip() == "4단지":
            return name
    for name in workbook.sheetnames:
        if "4단지" in str(name):
            return name
    return ""


def analyze_4danji_sheet(ws) -> Dict[str, object]:
    max_col = ws.max_column
    max_row = ws.max_row

    item_col = 4
    category_col = 3
    use_col = 6
    store_col = 7
    total_col = 8
    header_row = 4
    day_row = 5

    for r in range(1, min(max_row, 20) + 1):
        row_texts = [cell_text(ws, r, c) for c in range(1, min(max_col, 80) + 1)]
        if not any("품목명" in t for t in row_texts):
            continue

        header_row = r
        day_row = min(r + 1, max_row)
        for c, text in enumerate(row_texts, start=1):
            if not text:
                continue
            if text == "분류" or "분류" in text:
                category_col = c
            if "품목명" in text:
                item_col = c
            if "사용중인" in text:
                use_col = c
            if "재고수량" in text:
                store_col = c
            if "총" in text and "보유" in text:
                total_col = c
        break

    item_row_by_name: Dict[str, int] = {}
    item_rows: List[Dict[str, object]] = []
    blank_count = 0
    data_start_row = min(header_row + 2, max_row)

    for r in range(data_start_row, max_row + 1):
        item = cell_text(ws, r, item_col)
        if not item:
            blank_count += 1
            if blank_count >= 25:
                break
            continue
        blank_count = 0

        key = normalize_text(item)
        if not key or key in item_row_by_name:
            continue

        item_row_by_name[key] = r
        item_rows.append(
            {
                "row": r,
                "item": item,
                "item_key": key,
                "category": cell_text(ws, r, category_col),
            }
        )

    month_day_to_col: Dict[str, int] = {}
    for c in range(1, max_col + 1):
        header = cell_text(ws, header_row, c)
        if not header or "사용" not in header or "수량" not in header:
            continue

        month_match = re.search(r"(\d+)\s*월", header)
        if not month_match:
            continue
        month = int(month_match.group(1))

        for cc in range(c, max_col + 1):
            marker = cell_text(ws, day_row, cc)
            if not marker:
                continue
            if marker == "계":
                break
            try:
                day = int(marker)
            except ValueError:
                continue
            if 1 <= day <= 31:
                month_day_to_col[f"{month}-{day}"] = cc

    return {
        "item_row_by_name": item_row_by_name,
        "item_rows": item_rows,
        "month_day_to_col": month_day_to_col,
    }


def parse_date_parts(value: object) -> Optional[Tuple[int, int]]:
    text = str(value or "")[:10]
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", text)
    if not match:
        return None
    return int(match.group(2)), int(match.group(3))


def similarity_score(left: str, right: str) -> float:
    a = normalize_text(left)
    b = normalize_text(right)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0

    if a in b or b in a:
        short_len = min(len(a), len(b))
        long_len = max(len(a), len(b))
        ratio = short_len / long_len if long_len else 0.0
        if short_len <= 1:
            return ratio * 0.7
        if short_len == 2:
            return max(0.78, ratio)
        return max(0.86, ratio)

    score = SequenceMatcher(None, a, b).ratio()
    if a[0] == b[0]:
        score = min(1.0, score + 0.03)
    return max(0.0, min(1.0, score))


def find_best_name_match(item: str, candidates: List[str], min_score: float = 0.8, min_gap: float = 0.05):
    best = None
    second = None
    for idx, candidate in enumerate(candidates):
        score = similarity_score(item, candidate)
        candidate_info = {"index": idx, "name": candidate, "score": score}
        if best is None or score > best["score"]:
            second = best
            best = candidate_info
        elif second is None or score > second["score"]:
            second = candidate_info

    if best is None or best["score"] < min_score:
        return None
    if second is not None and best["score"] - second["score"] < min_gap:
        return None
    return best


def find_item_row_match(item: str, analyzed: Dict[str, object]):
    item_key = normalize_text(item)
    row_map = analyzed["item_row_by_name"]
    if item_key and item_key in row_map:
        return {
            "row": row_map[item_key],
            "item_key": item_key,
            "item": item,
            "matched_by": "exact",
            "score": 1.0,
        }

    item_rows = analyzed["item_rows"]
    best = find_best_name_match(item, [row["item"] for row in item_rows], min_score=0.8, min_gap=0.05)
    if not best:
        return None

    info = item_rows[best["index"]]
    return {
        "row": info["row"],
        "item_key": info["item_key"],
        "item": info["item"],
        "matched_by": "fuzzy",
        "score": best["score"],
    }


def resolve_zone_by_rack_barcode(rack_barcode: object) -> str:
    text = normalize_text(rack_barcode)
    if "guesthouse" in text or "guest-house" in text:
        return "게스트하우스"
    if "b2stairs" in text or "stairs" in text or "계단" in text:
        return "계단실"
    return "피트니스"


def apply_zone_row_visibility(ws, analyzed: Dict[str, object], allowed_item_keys: set) -> None:
    for info in analyzed["item_rows"]:
        row = int(info["row"])
        visible = info["item_key"] in allowed_item_keys
        ws.row_dimensions[row].hidden = not visible


def apply_outbound_to_sheet(ws, analyzed: Dict[str, object], transactions: List[Dict[str, object]]) -> Dict[str, object]:
    outbound = [tx for tx in transactions if str(tx.get("type", "")) == "출고" and not bool(tx.get("isInitial"))]
    if not outbound:
        raise ValueError("반영할 출고 기록이 없습니다.")

    applied_count = 0
    fuzzy_matched_count = 0
    missing_items = set()
    missing_dates = set()

    for tx in outbound:
        parts = parse_date_parts(tx.get("date", ""))
        if not parts:
            continue
        month, day = parts
        col = analyzed["month_day_to_col"].get(f"{month}-{day}")
        if not col:
            missing_dates.add(str(tx.get("date", "")))
            continue

        row_match = find_item_row_match(str(tx.get("item", "")), analyzed)
        if not row_match:
            missing_items.add(str(tx.get("item", "")))
            continue

        row = int(row_match["row"])
        qty = to_number(tx.get("qty", 0)) or 0.0
        current = to_number(ws.cell(row=row, column=col).value) or 0.0
        next_value = current + qty
        if abs(next_value - round(next_value)) < 1e-9:
            next_value = int(round(next_value))
        ws.cell(row=row, column=col).value = next_value

        applied_count += 1
        if row_match["matched_by"] == "fuzzy":
            fuzzy_matched_count += 1

    if applied_count == 0:
        raise ValueError("반영된 항목이 없습니다. 품목명 또는 날짜 칸을 확인해 주세요.")

    return {
        "applied_count": applied_count,
        "fuzzy_matched_count": fuzzy_matched_count,
        "missing_items_count": len(missing_items),
        "missing_dates_count": len(missing_dates),
        "outbound_count": len(outbound),
    }


def apply_outbound_by_zones(
    workbook,
    source_ws,
    analyzed: Dict[str, object],
    transactions: List[Dict[str, object]],
    baseline_rows: List[Dict[str, object]],
) -> Dict[str, object]:
    outbound = [tx for tx in transactions if str(tx.get("type", "")) == "출고" and not bool(tx.get("isInitial"))]

    zone_by_item_key: Dict[str, str] = {}

    for row in baseline_rows:
        item_key = normalize_text(row.get("item", ""))
        rack_barcode = str(row.get("rackBarcode", "")).strip()
        if not item_key or not rack_barcode or item_key in zone_by_item_key:
            continue
        zone_by_item_key[item_key] = resolve_zone_by_rack_barcode(rack_barcode)

    for tx in transactions:
        item_key = normalize_text(tx.get("item", ""))
        rack_barcode = str(tx.get("rackBarcode", "")).strip()
        if not item_key or not rack_barcode or item_key in zone_by_item_key:
            continue
        zone_by_item_key[item_key] = resolve_zone_by_rack_barcode(rack_barcode)

    for zone in EXPORT_ZONES:
        sheet_name = f"4단지_{zone}"
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    if "구역별_미반영" in workbook.sheetnames:
        del workbook["구역별_미반영"]

    unmatched_logs: List[List[object]] = []
    zone_summaries: List[Dict[str, object]] = []
    for zone in EXPORT_ZONES:
        zone_ws = workbook.copy_worksheet(source_ws)
        zone_ws.title = f"4단지_{zone}"

        allowed_item_keys = set(
            [
                info["item_key"]
                for info in analyzed["item_rows"]
                if zone_by_item_key.get(info["item_key"]) == zone
            ]
        )
        included_line_count = len(allowed_item_keys)

        apply_zone_row_visibility(zone_ws, analyzed, allowed_item_keys)

        applied_count = 0
        fuzzy_matched_count = 0

        for tx in outbound:
            item = str(tx.get("item", ""))
            item_key = normalize_text(item)
            row_match = find_item_row_match(item, analyzed)
            matched_key = row_match["item_key"] if row_match else item_key
            rack_barcode = str(tx.get("rackBarcode", "")).strip()
            tx_zone = (
                zone_by_item_key.get(matched_key)
                or zone_by_item_key.get(item_key)
            )
            if not tx_zone and rack_barcode:
                tx_zone = resolve_zone_by_rack_barcode(rack_barcode)
            if not tx_zone:
                if zone == "피트니스":
                    unmatched_logs.append([zone, item, tx.get("date", ""), tx.get("qty", 0), "렉바코드 없음(구역미분류)"])
                continue
            if tx_zone != zone:
                continue

            parts = parse_date_parts(tx.get("date", ""))
            if not parts:
                unmatched_logs.append([zone, item, tx.get("date", ""), tx.get("qty", 0), "날짜형식 오류"])
                continue
            month, day = parts
            col = analyzed["month_day_to_col"].get(f"{month}-{day}")
            if not col:
                unmatched_logs.append([zone, item, tx.get("date", ""), tx.get("qty", 0), "해당 날짜 칸 없음"])
                continue

            if not row_match or matched_key not in allowed_item_keys:
                unmatched_logs.append([zone, item, tx.get("date", ""), tx.get("qty", 0), "시트 품목 불일치"])
                continue

            row = int(row_match["row"])
            qty = to_number(tx.get("qty", 0)) or 0.0
            current = to_number(zone_ws.cell(row=row, column=col).value) or 0.0
            next_value = current + qty
            if abs(next_value - round(next_value)) < 1e-9:
                next_value = int(round(next_value))
            zone_ws.cell(row=row, column=col).value = next_value

            applied_count += 1
            if row_match["matched_by"] == "fuzzy":
                fuzzy_matched_count += 1

        zone_summaries.append(
            {
                "zone": zone,
                "included_line_count": included_line_count,
                "applied_count": applied_count,
                "fuzzy_matched_count": fuzzy_matched_count,
            }
        )

    if unmatched_logs:
        log_ws = workbook.create_sheet("구역별_미반영")
        log_ws.append(["구역", "품목명", "날짜", "수량", "사유"])
        for row in unmatched_logs:
            log_ws.append(row)

    return {
        "outbound_count": len(outbound),
        "unmatched_count": len(unmatched_logs),
        "zone_summaries": zone_summaries,
    }


def main():
    try:
        payload = json.loads(sys.stdin.read() or "{}")
        template_b64 = payload.get("template_b64", "")
        mode = str(payload.get("mode", "outbound"))
        file_name = payload.get("file_name", "4단지_출고반영.xlsx")
        transactions = payload.get("transactions", [])
        baseline_rows = payload.get("baseline_rows", [])

        if not template_b64:
            raise ValueError("기준표 템플릿 데이터가 없습니다.")
        if mode not in {"outbound", "zones"}:
            raise ValueError("지원하지 않는 내보내기 모드입니다.")
        if not isinstance(transactions, list):
            raise ValueError("거래 데이터 형식이 올바르지 않습니다.")
        if mode == "zones" and not isinstance(baseline_rows, list):
            raise ValueError("기준재고 데이터 형식이 올바르지 않습니다.")

        template_bytes = base64.b64decode(template_b64)
        workbook = load_workbook(io.BytesIO(template_bytes))
        sheet_name = find_4danji_sheet_name(workbook)
        if not sheet_name:
            raise ValueError("4단지 시트를 찾지 못했습니다.")

        ws = workbook[sheet_name]
        analyzed = analyze_4danji_sheet(ws)
        if not analyzed["item_rows"]:
            raise ValueError("4단지 시트에서 품목 데이터를 읽지 못했습니다.")

        if mode == "zones":
            summary = apply_outbound_by_zones(workbook, ws, analyzed, transactions, baseline_rows)
        else:
            summary = apply_outbound_to_sheet(ws, analyzed, transactions)

        out_buffer = io.BytesIO()
        workbook.save(out_buffer)
        out_b64 = base64.b64encode(out_buffer.getvalue()).decode("ascii")

        sys.stdout.write(
            json.dumps(
                {
                    "ok": True,
                    "file_name": file_name,
                    "workbook_b64": out_b64,
                    "summary": summary,
                },
                ensure_ascii=False,
            )
        )
    except Exception as exc:
        sys.stdout.write(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False))
        sys.exit(1)


if __name__ == "__main__":
    main()
